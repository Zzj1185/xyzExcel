"""
三坐标测量数据生成工具 - 网页版
Flask Web Application - 支持图片上传OCR识别

启动方法：
    python web_app.py

然后访问：
    http://localhost:5000
"""

from flask import Flask, render_template, request, jsonify, send_file
import random
import re
import io
import os
import base64
import urllib.request
import urllib.parse
import json
from datetime import datetime
from copy import copy

# Excel处理
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("警告: openpyxl未安装，Excel功能不可用。请运行: pip install openpyxl")

# Excel模板路径
EXCEL_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'model.xlsx')

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 最大16MB

# ========== 百度 OCR 配置 ==========
# 请在百度AI开放平台申请: https://console.bce.baidu.com/ai/#/ai/ocr/overview/index
BAIDU_API_KEY = 'bHq1UFwjFAeimASHLP3xXxBh'  # 填入你的 API Key
BAIDU_SECRET_KEY = 'UyhQrOEjHkJEBinnmrnxa27QfNO5SuLg'  # 填入你的 Secret Key
# ==================================

# 备用: OCR.space 免费API
OCR_SPACE_API_KEY = 'K85551736788957'
OCR_SPACE_API_URL = 'https://api.ocr.space/parse/image'


def get_baidu_access_token():
    """获取百度API访问令牌"""
    if not BAIDU_API_KEY or not BAIDU_SECRET_KEY:
        return None

    url = f"https://aip.baidubce.com/oauth/2.0/token?grant_type=client_credentials&client_id={BAIDU_API_KEY}&client_secret={BAIDU_SECRET_KEY}"

    try:
        req = urllib.request.Request(url, method='POST')
        with urllib.request.urlopen(req, timeout=10) as response:
            result = json.loads(response.read().decode('utf-8'))
            return result.get('access_token')
    except Exception as e:
        print(f"获取百度token失败: {e}")
        return None


def ocr_baidu(image_data):
    """使用百度OCR识别图片"""
    access_token = get_baidu_access_token()
    if not access_token:
        return None, "百度OCR未配置或获取token失败"

    url = f"https://aip.baidubce.com/rest/2.0/ocr/v1/accurate?access_token={access_token}"

    try:
        # Base64编码
        if isinstance(image_data, bytes):
            base64_image = base64.b64encode(image_data).decode('utf-8')
        else:
            base64_image = image_data

        # 发送请求
        payload = urllib.parse.urlencode({'image': base64_image}).encode('utf-8')
        req = urllib.request.Request(url, data=payload)
        req.add_header('Content-Type', 'application/x-www-form-urlencoded')

        with urllib.request.urlopen(req, timeout=30) as response:
            result = json.loads(response.read().decode('utf-8'))

        if 'error_code' in result:
            return None, f"百度OCR错误: {result.get('error_msg', '未知错误')}"

        # 返回完整结果列表
        return result.get('words_result', []), None

    except Exception as e:
        return None, f"百度OCR请求失败: {str(e)}"


def ocr_space(image_data, filename='image.png'):
    """使用OCR.space API识别图片中的文字（备用）"""
    try:
        # Base64编码图片
        if isinstance(image_data, bytes):
            base64_image = base64.b64encode(image_data).decode('utf-8')
        else:
            base64_image = image_data

        # 确定文件类型
        ext = filename.split('.')[-1].lower()
        if ext in ['jpg', 'jpeg']:
            mime = 'image/jpeg'
        elif ext == 'png':
            mime = 'image/png'
        else:
            mime = 'image/png'

        # 构建请求
        payload = {
            'apikey': OCR_SPACE_API_KEY,
            'base64Image': f'data:{mime};base64,{base64_image}',
            'language': 'eng',
            'isOverlayRequired': 'false',
            'detectOrientation': 'true',
            'scale': 'true',
            'OCREngine': '2',
        }

        # 发送请求
        data = urllib.parse.urlencode(payload).encode('utf-8')
        req = urllib.request.Request(OCR_SPACE_API_URL, data=data)
        req.add_header('Content-Type', 'application/x-www-form-urlencoded')

        with urllib.request.urlopen(req, timeout=30) as response:
            result = json.loads(response.read().decode('utf-8'))

        if result.get('IsErroredOnProcessing'):
            return None, result.get('ErrorMessage', ['OCR处理失败'])[0]

        parsed_results = result.get('ParsedResults', [])
        if parsed_results:
            text = parsed_results[0].get('ParsedText', '')
            return text, None

        return None, '未能识别到文字'

    except Exception as e:
        return None, f'OCR请求失败: {str(e)}'


def ocr_from_image(image_data, filename='image.png'):
    """智能OCR识别 - 优先使用百度，失败则用OCR.space"""
    # 先尝试百度OCR（更准确）
    if BAIDU_API_KEY and BAIDU_SECRET_KEY:
        text, error = ocr_baidu(image_data)
        if text:
            return text, None, 'baidu'
        print(f"百度OCR失败: {error}，尝试备用方案...")

    # 备用：OCR.space
    text, error = ocr_space(image_data, filename)
    if text:
        return text, None, 'ocr.space'

    return None, error or '所有OCR方案都失败了', None


def process_spatial_ocr(words_result):
    """
    基于位置信息对OCR结果进行行重组
    解决多列布局或位置错乱导致的解析问题
    """
    if not words_result:
        return ""

    # 添加辅助属性：中心点Y坐标
    items = []
    for item in words_result:
        location = item.get('location', {})
        top = location.get('top', 0)
        height = location.get('height', 0)
        left = location.get('left', 0)

        items.append({
            'text': item.get('words', ''),
            'top': top,
            'left': left,
            'height': height,
            'center_y': top + height / 2
        })

    # 按 top 排序
    items.sort(key=lambda x: x['top'])

    # 分行算法
    rows = []
    current_row = []

    if items:
        current_row = [items[0]]

        for item in items[1:]:
            last_item = current_row[-1]

            # 判断是否在同一行
            # 1. 垂直重叠 > 50%
            # 2. 或者中心点Y距离很小

            overlap = min(last_item['top'] + last_item['height'], item['top'] + item['height']) - max(last_item['top'], item['top'])
            min_height = min(last_item['height'], item['height'])

            is_same_row = False
            if min_height > 0 and overlap > min_height * 0.5:
                is_same_row = True
            elif abs(item['center_y'] - last_item['center_y']) < min_height * 0.5:
                is_same_row = True

            if is_same_row:
                current_row.append(item)
            else:
                rows.append(current_row)
                current_row = [item]

        if current_row:
            rows.append(current_row)

    # 对每一行按 left 排序并连接
    lines = []
    for row in rows:
        row.sort(key=lambda x: x['left'])
        line_text = " ".join([item['text'] for item in row])
        lines.append(line_text)

    return "\n".join(lines)


def parse_ocr_text(data):
    """从OCR识别的文本中解析坐标数据

    支持的格式：
    1. 百度OCR格式（每行一个值）：
       x=-201.865
       y=233.505
       z=108.338
       no.12
    2. 单行格式：x=77.463 y=33.497 z=50.000 no.1
    """
    points = []

    # 处理输入数据
    if isinstance(data, list):
        # 百度OCR返回的结构化数据，使用空间算法重组
        text = process_spatial_ocr(data)
    else:
        text = str(data)

    # 清理文本
    text = text.replace('\r\n', '\n').replace('\r', '\n')

    # 策略1：优先尝试按行解析（更精准，抗干扰，能解决错位问题）
    # 如果通过空间重组后，X,Y,Z在同一行，这里就能完美解析
    line_points = parse_coordinates_text(text)
    if len(line_points) > 0:
        return line_points

    # 策略2：全局正则匹配（兜底方案）
    # 适用于 X,Y,Z 分行显示的情况
    # 提取所有x、y、z值和编号
    x_values = re.findall(r'[xX]\s*[=:]\s*([-]?\d+[.,]\d+)', text)
    y_values = re.findall(r'[yY]\s*[=:]\s*([-]?\d+[.,]\d+)', text)
    z_values = re.findall(r'[zZ]\s*[=:]\s*([-]?\d+[.,]\d+)', text)
    no_values = re.findall(r'[nN][oO]\.?\s*(\d+)', text)

    # 找出数量最少的作为基准（避免重复匹配导致数量不一致）
    min_count = min(len(x_values), len(y_values), len(z_values))

    if min_count > 0:
        # 按顺序配对：每组x,y,z对应一个测量点
        for i in range(min_count):
            try:
                x = float(x_values[i].replace(',', '.'))
                y = float(y_values[i].replace(',', '.'))
                z = float(z_values[i].replace(',', '.'))

                # 尝试获取对应的编号，如果没有则使用序号
                if i < len(no_values):
                    point_id = int(no_values[i])
                else:
                    point_id = i + 1

                point = {
                    'id': point_id,
                    'x': x,
                    'y': y,
                    'z': z
                }

                # 检查是否已存在相同id的点
                if not any(p['id'] == point['id'] for p in points):
                    points.append(point)

            except (ValueError, IndexError):
                continue

    # 如果上面没匹配到，尝试按no.X分块解析
    if not points:
        blocks = re.split(r'(?=[nN][oO]\.?\s*\d+)', text)
        for block in blocks:
            point = parse_coordinate_block(block, len(points) + 1)
            if point:
                points.append(point)

    # 按编号排序
    points.sort(key=lambda x: x['id'])
    return points


def parse_coordinate_block(block, default_id):
    """解析单个坐标块"""
    # 提取编号
    no_match = re.search(r'[nN][oO]\.?\s*(\d+)', block)
    point_id = int(no_match.group(1)) if no_match else default_id

    # 提取x, y, z值
    x_match = re.search(r'[xX]\s*[=:]\s*([-]?\d+[.,]\d+)', block)
    y_match = re.search(r'[yY]\s*[=:]\s*([-]?\d+[.,]\d+)', block)
    z_match = re.search(r'[zZ]\s*[=:]\s*([-]?\d+[.,]\d+)', block)

    if x_match and y_match and z_match:
        try:
            return {
                'id': point_id,
                'x': float(x_match.group(1).replace(',', '.')),
                'y': float(y_match.group(1).replace(',', '.')),
                'z': float(z_match.group(1).replace(',', '.'))
            }
        except ValueError:
            pass

    return None


def parse_coordinates_text(text):
    """从文本中解析坐标数据"""
    points = []
    lines = text.strip().split('\n')

    for i, line in enumerate(lines):
        line = line.strip()
        if not line:
            continue

        point = parse_single_line(line, len(points) + 1)
        if point:
            points.append(point)

    return points


def parse_single_line(line, default_id):
    """解析单行坐标输入"""
    line = line.strip()

    # 格式1: 1, -77.463, 33.497, 50.000
    match = re.match(r'(\d+)\s*,\s*([-\d.]+)\s*,\s*([-\d.]+)\s*,\s*([-\d.]+)', line)
    if match:
        return {
            'id': int(match.group(1)),
            'x': float(match.group(2)),
            'y': float(match.group(3)),
            'z': float(match.group(4))
        }

    # 格式2: no.1 x=-77.463 y=33.497 z=50.000
    match = re.match(r'[nN][oO]\.?\s*(\d+)\s*[xX]\s*[=:]\s*([-\d.]+)\s*[yY]\s*[=:]\s*([-\d.]+)\s*[zZ]\s*[=:]\s*([-\d.]+)', line)
    if match:
        return {
            'id': int(match.group(1)),
            'x': float(match.group(2)),
            'y': float(match.group(3)),
            'z': float(match.group(4))
        }

    # 格式3: x=-77.463 y=33.497 z=50.000 no.1
    match = re.match(r'[xX]\s*[=:]\s*([-\d.]+)\s*[yY]\s*[=:]\s*([-\d.]+)\s*[zZ]\s*[=:]\s*([-\d.]+)\s*[nN][oO]\.?\s*(\d+)', line)
    if match:
        return {
            'id': int(match.group(4)),
            'x': float(match.group(1)),
            'y': float(match.group(2)),
            'z': float(match.group(3))
        }

    # 格式4: x=-77.463 y=33.497 z=50.000 (无编号)
    match = re.match(r'[xX]\s*[=:]\s*([-\d.]+)\s*[yY]\s*[=:]\s*([-\d.]+)\s*[zZ]\s*[=:]\s*([-\d.]+)', line)
    if match:
        return {
            'id': default_id,
            'x': float(match.group(1)),
            'y': float(match.group(2)),
            'z': float(match.group(3))
        }

    # 格式5: -77.463 33.497 50.000 (空格分隔)
    parts = line.replace(',', ' ').split()
    if len(parts) >= 3:
        try:
            return {
                'id': default_id,
                'x': float(parts[0]),
                'y': float(parts[1]),
                'z': float(parts[2])
            }
        except ValueError:
            pass

    return None


def generate_measurement_data(points, tolerance=0.03):
    """生成带随机偏差的测量数据"""
    measurements = []

    for point in points:
        for axis in ['X', 'Y', 'Z']:
            nominal = point[axis.lower()]
            deviation = random.uniform(-tolerance, tolerance)
            measured = nominal + deviation

            measurements.append({
                'point': point['id'],
                'axis': axis,
                'tol_min': -tolerance,
                'tol_max': tolerance,
                'nominal': nominal,
                'measured': measured,
                'deviation': deviation,
                'status': 'OK'
            })

    return measurements


def format_csv(measurements):
    """格式化为CSV字符串"""
    lines = ["测量点,轴,TOL,TOL,NomiNal,Measured,Dev/Mean,Error/OK"]

    current_point = None
    for m in measurements:
        point_str = f"no.{m['point']}" if m['point'] != current_point else ""
        current_point = m['point']

        line = f"{point_str},{m['axis']},{m['tol_min']:.2f},{m['tol_max']:.2f},{m['nominal']:.3f},{m['measured']:.3f},{m['deviation']:.3f},{m['status']}"
        lines.append(line)

    return '\n'.join(lines)


@app.route('/')
def index():
    """主页"""
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_image():
    """上传图片并OCR识别"""
    try:
        if 'image' not in request.files:
            return jsonify({'error': '请选择图片文件'}), 400

        file = request.files['image']
        if file.filename == '':
            return jsonify({'error': '请选择图片文件'}), 400

        # 读取图片数据
        image_data = file.read()

        # OCR识别
        ocr_text, error, ocr_provider = ocr_from_image(image_data, file.filename)

        if error:
            return jsonify({'error': error}), 400

        if not ocr_text:
            return jsonify({'error': '未能识别到文字'}), 400

        # 解析坐标
        points = parse_ocr_text(ocr_text)

        # 确保返回给前端的是文本字符串
        ocr_text_str = ocr_text
        if isinstance(ocr_text, list):
            ocr_text_str = process_spatial_ocr(ocr_text)

        return jsonify({
            'success': True,
            'ocr_text': ocr_text_str,
            'ocr_provider': ocr_provider or 'unknown',
            'points': points,
            'points_count': len(points)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/generate', methods=['POST'])
def generate():
    """生成测量数据"""
    try:
        data = request.get_json()
        text = data.get('coordinates', '')
        tolerance = float(data.get('tolerance', 0.03))

        if not text.strip():
            return jsonify({'error': '请输入坐标数据'}), 400

        # 解析坐标
        points = parse_coordinates_text(text)

        if not points:
            return jsonify({'error': '无法解析坐标数据，请检查格式'}), 400

        # 生成测量数据
        measurements = generate_measurement_data(points, tolerance)

        # 生成CSV
        csv_content = format_csv(measurements)

        return jsonify({
            'success': True,
            'points_count': len(points),
            'points': points,
            'csv': csv_content
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/generate_from_points', methods=['POST'])
def generate_from_points():
    """从已解析的点生成测量数据"""
    try:
        data = request.get_json()
        points = data.get('points', [])
        tolerance = float(data.get('tolerance', 0.03))

        if not points:
            return jsonify({'error': '没有坐标数据'}), 400

        # 生成测量数据
        measurements = generate_measurement_data(points, tolerance)

        # 生成CSV
        csv_content = format_csv(measurements)

        return jsonify({
            'success': True,
            'points_count': len(points),
            'points': points,
            'csv': csv_content
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/download', methods=['POST'])
def download():
    """下载CSV文件"""
    try:
        data = request.get_json()
        csv_content = data.get('csv', '')

        if not csv_content:
            return jsonify({'error': '无数据可下载'}), 400

        # 创建内存文件
        buffer = io.BytesIO()
        buffer.write(csv_content.encode('utf-8-sig'))  # 使用BOM以支持Excel
        buffer.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"测量数据_{timestamp}.csv"

        return send_file(
            buffer,
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ========== Excel 模板处理功能 ==========

def copy_cell_format(source_cell, target_cell):
    """复制单元格格式"""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


def copy_row(ws, source_row, target_row):
    """复制整行（包括格式和公式）"""
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(source_row, col)
        target_cell = ws.cell(target_row, col)

        if source_cell.value is not None:
            if isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                formula = source_cell.value.replace(str(source_row), str(target_row))
                target_cell.value = formula
            else:
                target_cell.value = source_cell.value

        copy_cell_format(source_cell, target_cell)


def update_sheet_data(wb, sheet_name, points, tolerance=0.03, image_data=None, filename=None, full_data=None):
    """更新单个工作表的数据

    Args:
        full_data: 可选的完整测量数据列表，每项包含 pointId, axis, nominal, measured, deviation
                   如果提供此参数，将直接使用这些数据而不生成随机偏差
    """
    if sheet_name not in wb.sheetnames:
        return f"工作表 '{sheet_name}' 不存在于模板中"

    ws = wb[sheet_name]

    try:
        # 更新 B3 单元格：编号 + 后缀
        if filename:
            # 从文件名提取编号（第二个 - 之前的部分）
            # 例如 "P25-488-前门扶手本体上（左右）-三坐标报告" -> "P25-488"
            parts = filename.split('-')
            if len(parts) >= 2:
                prefix = f"{parts[0]}-{parts[1]}"
                # 根据工作表名确定后缀
                if sheet_name == '前模仁':
                    suffix = '-CAV1（A）'
                else:  # 后模仁
                    suffix = '-COR1（B）'
                ws.cell(3, 2).value = prefix + suffix

        # 生成或使用测量数据
        measurement_data = []

        if full_data and len(full_data) > 0:
            # 使用用户提供的完整数据
            for item in full_data:
                measurement_data.append({
                    'point': item.get('pointId', 1),
                    'axis': item.get('axis', 'X').upper(),
                    'nominal': float(item.get('nominal', 0)),
                    'measured': float(item.get('measured', 0))
                })
        else:
            # 生成随机偏差数据
            for p in points:
                for axis in ['X', 'Y', 'Z']:
                    nominal = float(p[axis.lower()])
                    deviation = random.uniform(-tolerance, tolerance)
                    measured = nominal + deviation
                    measurement_data.append({
                        'point': p['id'],
                        'axis': axis,
                        'nominal': nominal,
                        'measured': measured
                    })

        # 计算需要的测量点数量
        max_point = max(item['point'] for item in measurement_data)

        # 统计当前测量点数量
        current_points = sum(1 for row in range(1, ws.max_row + 1)
                             if ws.cell(row, 1).value and "测量点" in str(ws.cell(row, 1).value))

        # 添加测量点
        if max_point > current_points:
            for new_point in range(current_points + 1, max_point + 1):
                new_start_row = 13 + (new_point - 1) * 6
                for offset in range(5):
                    ws.insert_rows(new_start_row + offset)
                    copy_row(ws, 19 + offset, new_start_row + offset)
                ws.cell(new_start_row, 1).value = f"测量点{new_point}"

        # 删除多余测量点
        elif max_point < current_points:
            for point_to_delete in range(current_points, max_point, -1):
                start_row = 13 + (point_to_delete - 1) * 6
                ws.delete_rows(start_row, 5)

        # 更新测量数据
        for item in measurement_data:
            point, axis = item['point'], item['axis']
            nominal, measured = item['nominal'], item['measured']

            axis_offset = {'X': 0, 'Y': 1, 'Z': 2}[axis]
            target_row = 15 + (point - 1) * 6 + axis_offset

            # 更新数值
            ws.cell(target_row, 7, nominal)
            ws.cell(target_row, 8, measured)

            # 设置标准数字格式
            ws.cell(target_row, 7).number_format = "0.000"
            ws.cell(target_row, 8).number_format = "0.000"

            # 强制设置偏差列(I列)格式，确保负数显示为 -0.000 而不是红色括号
            ws.cell(target_row, 9).number_format = "0.000"

        # 插入图片
        if image_data:
            try:
                # 去掉base64头 (data:image/png;base64,...)
                if ',' in image_data:
                    image_data = image_data.split(',')[1]

                img_bytes = base64.b64decode(image_data)
                img_stream = io.BytesIO(img_bytes)
                img = ExcelImage(img_stream)

                # 寻找 A7 所在的合并单元格
                img_anchor = 'A7'
                merged_range = None

                # 遍历所有合并单元格范围
                for rng in ws.merged_cells.ranges:
                    # 检查A7 (Col 1, Row 7) 是否在范围内
                    if (rng.min_col <= 1 <= rng.max_col) and (rng.min_row <= 7 <= rng.max_row):
                        merged_range = rng
                        break

                if merged_range:
                    # 计算合并区域的总宽度和高度 (估算)
                    total_width_px = 0
                    for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                        col_letter = get_column_letter(col_idx)
                        col_w = ws.column_dimensions[col_letter].width
                        if col_w is None: col_w = 8.38 # 默认宽度
                        total_width_px += col_w * 7 # 粗略转换: 1 char width ≈ 7 pixels

                    total_height_px = 0
                    for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
                        row_h = ws.row_dimensions[row_idx].height
                        if row_h is None: row_h = 15 # 默认高度
                        total_height_px += row_h * 1.33 # 1 point ≈ 1.33 pixels

                    # 调整图片大小以适应区域 (留出10%边距)
                    target_w = max(total_width_px * 0.9, 100) # 至少100px
                    target_h = max(total_height_px * 0.9, 100) # 至少100px

                    if img.width > target_w or img.height > target_h:
                        scale_w = target_w / img.width if img.width > 0 else 1
                        scale_h = target_h / img.height if img.height > 0 else 1
                        scale = min(scale_w, scale_h)

                        img.width = int(img.width * scale)
                        img.height = int(img.height * scale)

                    # 水平居中 (精确计算偏移量)
                    remaining_w = total_width_px - img.width
                    offset_x = max(0, remaining_w / 2)

                    # 垂直居中
                    remaining_h = total_height_px - img.height
                    offset_y = max(0, remaining_h / 2)

                    # 寻找包含偏移起点的列
                    current_x = 0
                    anchor_col = merged_range.min_col
                    col_offset_px = 0

                    for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                        col_letter = get_column_letter(col_idx)
                        col_w = ws.column_dimensions[col_letter].width
                        if col_w is None: col_w = 8.38
                        px_w = col_w * 7

                        if current_x + px_w > offset_x:
                            anchor_col = col_idx
                            col_offset_px = offset_x - current_x
                            break

                        current_x += px_w

                    # 寻找包含偏移起点的行
                    current_y = 0
                    anchor_row = merged_range.min_row
                    row_offset_px = 0

                    for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
                        row_h = ws.row_dimensions[row_idx].height
                        if row_h is None: row_h = 15
                        px_h = row_h * 1.33

                        if current_y + px_h > offset_y:
                            anchor_row = row_idx
                            row_offset_px = offset_y - current_y
                            break

                        current_y += px_h

                    # 使用 OneCellAnchor 进行精确放置
                    # EMU (English Metric Unit) 转换: 1 pixel = 9525 EMUs
                    col_offset_emu = int(col_offset_px * 9525)
                    row_offset_emu = int(row_offset_px * 9525)

                    marker = AnchorMarker(col=anchor_col - 1, colOff=col_offset_emu, row=anchor_row - 1, rowOff=row_offset_emu)
                    size = XDRPositiveSize2D(cx=int(img.width * 9525), cy=int(img.height * 9525))
                    img.anchor = OneCellAnchor(_from=marker, ext=size)

                    ws.add_image(img)

                else:
                    # 如果不是合并单元格，使用之前的简单逻辑
                    max_width = 300
                    max_height = 200
                    if img.width > max_width or img.height > max_height:
                        scaling = min(max_width / img.width, max_height / img.height)
                        img.width = int(img.width * scaling)
                        img.height = int(img.height * scaling)

                    ws.add_image(img, 'A7')
            except Exception as e:
                print(f"插入图片失败: {e}")

        return None

    except Exception as e:
        return f"更新工作表 '{sheet_name}' 失败: {str(e)}"


def generate_excel_from_template(points, sheet_name, tolerance=0.03, image_data=None, sheets_data=None, filename=None):
    """从模板生成Excel文件 (支持单表或多表)"""
    if not HAS_OPENPYXL:
        return None, "openpyxl未安装，无法生成Excel"

    if not os.path.exists(EXCEL_TEMPLATE_PATH):
        return None, f"模板文件不存在: {EXCEL_TEMPLATE_PATH}"

    try:
        # 加载模板
        wb = openpyxl.load_workbook(EXCEL_TEMPLATE_PATH)

        # 处理多表数据
        if sheets_data:
            for item in sheets_data:
                full_data = item.get('full_data')  # 获取完整数据（如果有）
                err = update_sheet_data(
                    wb,
                    item['sheet_name'],
                    item['points'],
                    item.get('tolerance', tolerance),
                    item.get('image_data'),
                    filename,
                    full_data
                )
                if err:
                    return None, err
        else:
            # 兼容旧的单表模式
            err = update_sheet_data(wb, sheet_name, points, tolerance, image_data, filename)
            if err:
                return None, err

        # 保存到内存
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        wb.close()

        return buffer, None

    except Exception as e:
        return None, f"生成Excel失败: {str(e)}"


@app.route('/generate_excel', methods=['POST'])
def generate_excel():
    """生成Excel文件并下载"""
    try:
        data = request.get_json()

        # 检查是否为多表请求
        sheets_data = data.get('sheets_data')

        # 单表请求参数
        points = data.get('points', [])
        sheet_name = data.get('sheet_name', '前模仁')
        image_data = data.get('image_data')

        filename = data.get('filename', '')
        tolerance = float(data.get('tolerance', 0.03))

        if not sheets_data and not points:
            return jsonify({'error': '没有坐标数据'}), 400

        if not filename:
            filename = f"三坐标报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        # 确保文件名以.xlsx结尾
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        # 生成Excel (传入原始文件名用于提取编号)
        buffer, error = generate_excel_from_template(points, sheet_name, tolerance, image_data, sheets_data, filename)

        if error:
            return jsonify({'error': error}), 400

        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# 创建templates目录
templates_dir = os.path.join(os.path.dirname(__file__), 'templates')
os.makedirs(templates_dir, exist_ok=True)

if __name__ == '__main__':
    print("=" * 60)
    print("三坐标测量数据生成工具 - 网页版")
    print("=" * 60)
    print("\n启动服务器...")
    print("请在浏览器中访问: http://localhost:5000")
    print("\n按 Ctrl+C 停止服务器")
    print("=" * 60)

    app.run(host='0.0.0.0', port=5000, debug=True)
